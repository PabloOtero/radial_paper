# -*- coding: utf-8 -*-
"""
OPEN RADIALES NETCDF FILE AND MAKE SOME PLOTS

Example of how to open the netcdf file of the dataset: 
    
"Temperature, conductivity and salinity data of the RADIALES Program, 
Vigo, 1987-2020" 

and distributed at:
    
https://www.seanoe.org/data/00828/94008/


Created on Fri Mar  4 10:51:57 2022
@author: pablo.otero@ieo.csic.es
"""

import numpy as np
import xarray as xr
import matplotlib.pyplot as plt
import datetime
import pandas
import matplotlib.dates as mdates
import pandas as pd
import seaborn as sns; sns.set_theme()
from pathlib import Path
import gsw

cmap = plt.cm.RdYlBu_r

# Path for netcdf files (include here your relative path)
data_path = 'data\ctd_netcdf_qc'

# Input netcdf file
file = "radial_vigo_cf_odv_qc.nc"


#%% LOAD DATA

# Open the file
file_to_open = Path.cwd() / data_path / file
ds = xr.open_dataset(file_to_open)

# Get info
ds.info

# Get cruise names (from array of bytes to list of strings)
cruise = list(ds.variables['metavar1'].values)
cruise = [str(c,'utf-8') for c in cruise]

station = ds.variables['metavar2'].values
station_decoded = []
for station_coded in station:
    station_decoded.append(int(station_coded.decode('UTF-8').split('-')[0].replace('v', '')))
station_names = np.unique(station_decoded)
station_decoded = np.array(station_decoded)

longitude = ds.variables['longitude'].values
latitude = ds.variables['latitude'].values
time = ds.variables['date_time'].values

# press is stored as 'var7'
z=ds.variables['var7'].values

# the file contains 3 temperature variables
# a) temperature from a profile with only one mounted sensor
# b) temperature from the main (first) sensor when two sensors coexist
# c) temperature from the secondary sensor when two sensors coexist
# we merge a) and b) in a single array 

# The option a) is stored as 'var8'
t=ds.variables['var8'].values
t_flag = ds.variables['var8_qc'].values
t_flag[np.where(t_flag == b'1')] = 1
t_flag[np.where(t_flag == b'2')] = 2
t_flag[np.where(t_flag == b'3')] = 3
t_flag[np.where(t_flag == b'4')] = 4
t[np.where(t_flag == 3)] = np.nan
t[np.where(t_flag == 4)] = np.nan

# The option b) is stored as 'var2'
t1=ds.variables['var9'].values
t1_flag = ds.variables['var9_qc'].values
t1_flag[np.where(t1_flag == b'1')] = 1
t1_flag[np.where(t1_flag == b'2')] = 2
t1_flag[np.where(t1_flag == b'3')] = 3
t1_flag[np.where(t1_flag == b'4')] = 4
t1[np.where(t1_flag == 3)] = np.nan
t1[np.where(t1_flag == 4)] = np.nan

# The option c) is stored as 'var10'
t2=ds.variables['var10'].values
t2_flag = ds.variables['var9_qc'].values
t2_flag[np.where(t2_flag == b'1')] = 1
t2_flag[np.where(t2_flag == b'2')] = 2
t2_flag[np.where(t2_flag == b'3')] = 3
t2_flag[np.where(t2_flag == b'4')] = 4
t2[np.where(t2_flag == 3)] = np.nan
t2[np.where(t2_flag == 4)] = np.nan

# if not a), then check if b)
for indx, profile in enumerate(t):
    if np.all(np.isnan(profile)):
        print('kk')
        t[indx,:] = t2[indx,:]


# Salinity         
s=ds.variables['var6'].values
s_flag = ds.variables['var6_qc'].values
s_flag[np.where(s_flag == b'1')] = 1
s_flag[np.where(s_flag == b'2')] = 2
s_flag[np.where(s_flag == b'3')] = 3
s_flag[np.where(s_flag == b'4')] = 4
s[np.where(s_flag == 3)] = np.nan
s[np.where(s_flag == 4)] = np.nan

ds.close()


#%% Fill gaps in depth only if exists previous and consecutive values
for ncast, profile in enumerate(t):
    for indx, element in enumerate(profile):
        if (indx > 1) and (indx < len(profile)-1):
            if( (np.isnan(element)) and (np.isnan(profile[indx-1]) == False) and (np.isnan(profile[indx+1]) == False) ):
                profile[indx] = np.mean([profile[indx-1],profile[indx+1]])
                print('From NaN to ', np.mean([profile[indx-1],profile[indx+1]]), ' at ', str(indx+1), 'm depth' )
    t[ncast,:] = profile


for ncast, profile in enumerate(s):
    for indx, element in enumerate(profile):
        if (indx > 1) and (indx < len(profile)-1):
            if( (np.isnan(element)) and (np.isnan(profile[indx-1]) == False) and (np.isnan(profile[indx+1]) == False) ):
                profile[indx] = np.mean([profile[indx-1],profile[indx+1]])
                print('From NaN to ', np.mean([profile[indx-1],profile[indx+1]]), ' at ', str(indx+1), 'm depth' )
    s[ncast,:] = profile


#%% Organize the data
# We create variables 'temp' and 'salt with dimensions (time, station, depth)

# Interpolate profiles to a constant depth
# max z can be found with np.nanmax(z) and is 144 dbar
z_new = np.arange(1, 150, 1)
t_new = []
s_new = []
# Create a 2D array with nans 
t_new = np.empty((len(t),len(z_new),))*np.nan
s_new = np.empty((len(t),len(z_new),))*np.nan
for idx, profile in enumerate(t):
    t_new[idx, :] = np.interp(z_new, z[idx,:], t[idx,:], left=np.nan, right=np.nan)
    s_new[idx, :] = np.interp(z_new, z[idx,:], s[idx,:], left=np.nan, right=np.nan)


#%% CREATE A PANDAS DATAFRAME including temp, salt and density

# Warning, TEOS-10 needs absolute salinity
SA = gsw.SA_from_SP(s_new, z_new, 42, -10.5)
dens_potential = gsw.pot_rho_t_exact(SA, t_new, z_new, 0)-1000

CT = gsw.CT_from_t(SA, t_new, z_new)
dens_insitu = gsw.rho(SA, CT, z_new)-1000


df = pd.DataFrame()
for idx in range(0,len(t_new)):
    time_equal_vector = np.full((len(z_new)), time[idx])
    station_equal_vector = np.full((len(z_new)), station_decoded[idx])
    df2 = pd.DataFrame({'time': list(time_equal_vector), 'station': list(station_equal_vector), 'depth': list(z_new), 'temp': list(t_new[idx,:]), 'salt': list(s_new[idx,:]), 'dens_potential': list(dens_potential[idx,:]), 'dens_insitu': list(dens_insitu[idx,:])}, columns=['time', 'station', 'depth', 'temp', 'salt', 'dens_potential', 'dens_insitu'])
    df = pd.concat([df, df2], axis=0)
df.set_index('time', inplace=True)
df.sort_index(inplace=True) 

        

#%% PLOT PANEL

labels = ['Temperature (\u00B0C)', 'Salinity', 'Density']
names = ['temp', 'salt', 'dens_insitu']

station_names = [15, 1, 3, 5]

locator = mdates.YearLocator()

surface_meters_to_mask = 3

for label, name in zip(labels, names):
    
    if name == 'salt':
        vmin = 33.8
        vmax = 36.0
        levels = np.arange(vmin, vmax, 0.2)        
    elif name == 'dens_insitu':
        vmin = 25.2
        vmax = 27.2
        levels = np.arange(vmin, vmax, 0.2)                       
    else:
        vmin = 11
        vmax = 22
        levels = np.arange(vmin, vmax, 1)  

    fig, axs = plt.subplots(4, figsize=(20, 10), sharex=True)    
    
    # Set the ticks and ticklabels for all axes
    plt.setp(axs, yticks=np.arange(-150, 25, 25))
    
    
    for stat, ax in zip(station_names, axs.ravel()):
    
        pv = df[df['station']==stat].pivot(columns='depth', values=name)
        pv = pv.resample('M').mean()
        
        # Create a new dataframe and mask first meters to avoid pcolormesh artifacts
        pv2 = pv.copy() 
        pv2.iloc[:,0:surface_meters_to_mask]=np.nan
      
        cp = ax.pcolormesh(pv2.index.to_pydatetime(), -pv2.columns, pv2.values.T, cmap=cmap, zorder=1, vmin=vmin, vmax=vmax)

        cf = ax.contourf(pv2.index.to_pydatetime(), -pv2.columns, pv2.values.T, levels, cmap=cmap, zorder=2, extend='both')        
        
        ax.grid(True)
        ax.set_ylim([-150, 0])
        ylim = -(ax.get_ylim()[0])
        ax.xaxis.set_major_locator(locator)
        minyear = 1987
        mintime = datetime.datetime(year=minyear, month=1, day=1)
        maxyear = 2020
        maxtime = datetime.datetime(year=maxyear, month=12, day=31)        
        
        ax.set_xlim( mintime, maxtime )  
        ax.set_xlabel("Year", fontsize=10, loc="left")
        ax.set_ylabel("Depth (m)", fontsize=10, loc="top")
        ax.tick_params(axis='both', which='major', labelsize=9)
        
        cb = plt.colorbar(cf, ax=ax, label=label, ticks=levels)
        locator = mdates.YearLocator()

        
    title = 'Fig_timeseries_panel' + '_' + name + '.png'
    plt.savefig(title, bbox_inches='tight', pad_inches=.2, dpi=600)
    plt.close(fig)


#%% PLOT CHRONOGRAM WITH SURVEYS

df = pd.DataFrame(data = time, columns = ['time'])
df = df.drop_duplicates()
df['day'] = df['time'].dt.day
#df['month'] = df['time'].dt.month_name().str[:3]
df['month'] = df['time'].dt.month
#df['year'] = df['time'].dt.year
df['year'] = df['time'].dt.strftime('%Y')
df.drop('time', inplace=True, axis=1)
df = df.drop_duplicates()
df['count'] = 1

# create pivot table, days will be columns, hours will be rows
# if you want to compute the total number of surveys per month...sum events
piv = pd.pivot_table(df,index='month', columns='year', values='count', fill_value=0, aggfunc=np.sum)
# or just keep only the first
# piv = pd.pivot_table(df,index='month', columns='year', values='count', fill_value=0, aggfunc='first')
piv.index=['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
piv = piv.T

piv_dates = pd.pivot_table(df,index='month', columns='year', values='day', fill_value=0, aggfunc='first')
piv_dates = piv_dates.T

#plot pivot table as heatmap
fig, ax = plt.subplots(figsize=(6,9))
im = ax.imshow(piv, origin="upper", cmap=plt.cm.get_cmap('Blues', 3))

# Loop over data dimensions and create text annotations.
for i in range(piv_dates.shape[0]):
    for j in range(piv_dates.shape[1]):
        text = ax.text(j, i, piv_dates.iloc[i, j], ha="center", va="center", color="w")

fig.colorbar(im, ax=ax, label="Number of surveys", fraction=0.014, pad=0.01)
ax.set_xticks(np.arange(len(piv.columns)))
ax.set_yticks(np.arange(len(piv.index)))
ax.set_xticklabels(piv.columns, rotation=90)
ax.set_yticklabels(piv.index)
ax.set_xlabel("Month", fontsize=12, loc="left")
ax.set_ylabel("Year", fontsize=11, loc="top")   

# Minor ticks
ax.set_xticks(np.arange(-.5, len(piv.columns)-0.5, 1), minor=True)
ax.set_yticks(np.arange(-0.5, len(piv.index)-0.5, 1), minor=True)
ax.tick_params(axis='both', which='both', labelsize=11)

# Gridlines based on minor ticks
ax.grid(False)
ax.grid(which='minor', color='w', linestyle='-', linewidth=1.5)
ax.grid(which='major', linewidth=0)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.spines['bottom'].set_visible(False)
ax.spines['left'].set_visible(False)
fig.tight_layout()
plt.show()

title = 'Fig1_colormap_surveys_simple.png'
plt.savefig(title, bbox_inches='tight', pad_inches=.2, dpi=600)
plt.close(fig)




#%% SAVE ALL DATA TO EXCEL (with more than 65K rows)

from openpyxl import load_workbook

filename = 'radiales_thermohaline.xlsx'

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False, 
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.
     
        Parameters:
          filename : File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
          df : dataframe to save to workbook
          sheet_name : Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]
     
        Returns: None
     
        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
     
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')
          
        try:
            writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
            # try to open an existing workbook
            writer.book = load_workbook(filename)
             
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
     
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
             
            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            writer = pd.ExcelWriter(filename, engine='openpyxl')
     
        if startrow is None:
            startrow = 0
     
        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
     
        # save the workbook
        writer.save()


append_df_to_excel(filename, df, sheet_name='Sheet1')
     